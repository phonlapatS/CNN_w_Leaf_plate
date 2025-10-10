# GUI.py
# -*- coding: utf-8 -*-
# Required: pip install customtkinter ultralytics opencv-python pillow firebase-admin openpyxl
# Put your Firebase service account JSON next to this file as: serviceAccountKey.json

import os, sys, time, json, uuid, glob, signal
from datetime import datetime, date, timedelta, timezone

import customtkinter as ctk
import tkinter as tk
from tkinter import ttk
from tkinter import font as tkfont
from tkinter import messagebox

import cv2
import numpy as np
from PIL import Image, ImageTk
from ultralytics import YOLO

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict

# -------- Emoji Support --------
try:
    import emojis
    EMOJI_AVAILABLE = True
except ImportError:
    EMOJI_AVAILABLE = False
    print("Warning: emojis library not available. Install with: pip install emojis")

try:
    from colorama import init, Fore, Back, Style
    init(autoreset=True)
    COLORAMA_AVAILABLE = True
except ImportError:
    COLORAMA_AVAILABLE = False
    print("Warning: colorama library not available. Install with: pip install colorama")

# -------- Firebase Admin SDK --------
import firebase_admin
from firebase_admin import credentials, db


# ================================
# Excel Viewer (รายวัน) – ใช้สเกลอัตโนมัติ
# ================================
class ExcelViewerDialog(ctk.CTkToplevel):
    """Excel viewer/editor (รายวัน): เลือกไฟล์จาก dropdown -> โหลดทันที; ดับเบิลคลิกเพื่อแก้ไข"""
    def __init__(self, parent, folder, title="Excel Viewer"):
        super().__init__(parent)
        self.parent = parent
        self.folder = folder
        self.title(title)

        scale = getattr(parent, "SCALE", 1.0)
        def px(v): return int(round(v * scale))

        self.geometry(f"{px(980)}x{px(560)}")
        self.resizable(True, True)
        self.grab_set()

        # Top bar
        top = ctk.CTkFrame(self, fg_color="transparent")
        top.pack(fill="x", padx=px(10), pady=(px(10), px(6)))

        ctk.CTkLabel(top, text="เลือกไฟล์:", font=ctk.CTkFont(size=px(14), weight="bold")).pack(side="left", padx=(0, px(8)))
        self.combo = ttk.Combobox(top, state="readonly", width=48)
        self.combo.pack(side="left")
        self.combo.bind("<<ComboboxSelected>>", lambda e: self._load_selected())

        ctk.CTkButton(top, text="รีเฟรช", width=px(80), command=self._refresh_file_list).pack(side="left", padx=px(8))

        # Table
        mid = ctk.CTkFrame(self)
        mid.pack(fill="both", expand=True, padx=px(10), pady=(0, px(10)))

        self.tree = ttk.Treeview(mid, show="headings")
        self.tree.pack(side="left", fill="both", expand=True)
        vsb = ttk.Scrollbar(mid, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(mid, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        self.tree.bind("<Double-1>", self._begin_edit)

        # Bottom
        bot = ctk.CTkFrame(self, fg_color="transparent")
        bot.pack(fill="x", padx=px(10), pady=(0, px(10)))
        ctk.CTkButton(bot, text="บันทึกทั้งหมดอีกครั้ง", command=self._save_all).pack(side="right", padx=px(6))
        ctk.CTkButton(bot, text="ปิด", command=self.destroy).pack(side="right")

        self.current_path = None
        self.headers = []
        self.header_row_idx = 1

        self._refresh_file_list()

    def _refresh_file_list(self):
        prev = self.combo.get()
        paths = sorted(
            glob.glob(os.path.join(self.folder, "Report_*.xlsx")),
            key=lambda p: os.path.getmtime(p),
            reverse=True,
        )
        names = [os.path.basename(p) for p in paths]
        self.combo["values"] = names

        if not names:
            self.combo.set("")
            self.current_path = None
            self.tree.delete(*self.tree.get_children())
            self.title("Excel Viewer - (ไม่มีไฟล์)")
            return

        if prev in names:
            self.combo.current(names.index(prev))
        else:
            self.combo.current(0)

        self._load_selected()

    def _load_selected(self):
        name = self.combo.get()
        if not name:
            return
        path = os.path.join(self.folder, name)
        if not os.path.exists(path):
            messagebox.showerror("Excel", f"ไม่พบไฟล์: {path}")
            return
        self.current_path = path
        self._load_excel_to_tree(path)

    def _detect_header_row(self, sh):
        for r in range(1, min(sh.max_row, 20) + 1):
            val = sh.cell(row=r, column=1).value
            if isinstance(val, str) and val.strip() == "วันที่":
                return r
        return 1

    def _load_excel_to_tree(self, path):
        wb = openpyxl.load_workbook(path, data_only=True)
        sh = wb.active

        self.header_row_idx = self._detect_header_row(sh)
        self.headers = [cell.value if cell.value is not None else "" for cell in sh[self.header_row_idx]]
        if not self.headers:
            self.headers = ["วันที่", "เวลา", "จานที่", "รหัสชุด", "รูปทรงจาน", "ตำหนิที่พบ", "หมายเหตุ"]

        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = [str(i) for i in range(len(self.headers))]
        for i, h in enumerate(self.headers):
            self.tree.heading(str(i), text=str(h))
            self.tree.column(str(i), width=140, anchor="center")

        for row in sh.iter_rows(min_row=self.header_row_idx + 1, values_only=True):
            vals = [("" if v is None else str(v)) for v in row]
            if len(vals) < len(self.headers):
                vals += [""] * (len(self.headers) - len(vals))
            elif len(vals) > len(self.headers):
                vals = vals[:len(self.headers)]
            if any(v != "" for v in vals):
                self.tree.insert("", "end", values=vals)

        wb.close()
        self.title(f"Excel Viewer - {os.path.basename(path)}")

    def _begin_edit(self, event):
        if not self.current_path:
            return
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if not row_id or not col_id:
            return
        col_idx = int(col_id.replace("#", "")) - 1
        x, y, w, h = self.tree.bbox(row_id, col_id)
        cur = self.tree.set(row_id, col_id)

        entry = tk.Entry(self.tree)
        entry.insert(0, cur)
        entry.select_range(0, tk.END)
        entry.focus()
        entry.place(x=x, y=y, width=w, height=h)

        def commit(_=None):
            new_val = entry.get()
            entry.destroy()
            values = list(self.tree.item(row_id, "values"))
            values[col_idx] = new_val
            self.tree.item(row_id, values=values)
            self._write_back_cell(row_id, col_idx, new_val)

        def cancel(_=None):
            entry.destroy()

        entry.bind("<Return>", commit)
        entry.bind("<Escape>", cancel)
        entry.bind("<FocusOut>", commit)

    def _tree_index_of(self, row_id):
        return self.tree.get_children().index(row_id)

    def _write_back_cell(self, row_id, col_idx, value):
        try:
            wb = openpyxl.load_workbook(self.current_path)
            sh = wb.active
            excel_row = self.header_row_idx + 1 + self._tree_index_of(row_id)
            excel_col = col_idx + 1
            sh.cell(row=excel_row, column=excel_col, value=value)
            wb.save(self.current_path)
            wb.close()
        except Exception as e:
            print(f"[Excel write-back] {e}")

    def _save_all(self):
        if not self.current_path:
            return
        try:
            wb = openpyxl.load_workbook(self.current_path)
            sh = wb.active
            r0 = self.header_row_idx + 1
            for i, row_id in enumerate(self.tree.get_children(), start=r0):
                vals = list(self.tree.item(row_id, "values"))
                for j, v in enumerate(vals, start=1):
                    sh.cell(row=i, column=j, value=v)
            wb.save(self.current_path)
            wb.close()
            messagebox.showinfo("Excel", "บันทึกเรียบร้อย")
        except Exception as e:
            messagebox.showerror("Excel", f"บันทึกไม่สำเร็จ:\n{e}")


# ================================
# Excel Viewer (รายสัปดาห์) – แถบหัววันสีเทา/ห้ามแก้ไข + สเกลอัตโนมัติ
# ================================
class WeeklyExcelViewerDialog(ctk.CTkToplevel):
    def __init__(self, parent, folder, title="Excel Viewer - Weekly"):
        super().__init__(parent)
        self.parent = parent
        self.folder = folder
        self.title(title)

        scale = getattr(parent, "SCALE", 1.0)
        def px(v): return int(round(v * scale))

        self.geometry(f"{px(980)}x{px(560)}")
        self.resizable(True, True)
        self.grab_set()

        top = ctk.CTkFrame(self, fg_color="transparent")
        top.pack(fill="x", padx=px(10), pady=(px(10), px(6)))

        ctk.CTkLabel(top, text="เลือกไฟล์:", font=ctk.CTkFont(size=px(14), weight="bold")).pack(side="left", padx=(0, px(8)))
        self.combo = ttk.Combobox(top, state="readonly", width=48)
        self.combo.pack(side="left")
        self.combo.bind("<<ComboboxSelected>>", lambda e: self._load_selected())

        ctk.CTkButton(top, text="รีเฟรช", width=px(80), command=self._refresh_file_list).pack(side="left", padx=px(8))

        mid = ctk.CTkFrame(self)
        mid.pack(fill="both", expand=True, padx=px(10), pady=(0, px(10)))

        self.tree = ttk.Treeview(mid, show="headings")
        self.tree.pack(side="left", fill="both", expand=True)
        vsb = ttk.Scrollbar(mid, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(mid, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        # หัววัน: เทาอ่อน + bold + ห้ามแก้ไข
        self.tree.tag_configure(
            "day_header",
            background="#E6E6E6",
            foreground="#333333",
            font=tkfont.Font(family="Arial", size=max(9, px(10)), weight="bold")
        )
        self.protected_rows = set()

        self.tree.bind("<Double-1>", self._begin_edit)

        bot = ctk.CTkFrame(self, fg_color="transparent")
        bot.pack(fill="x", padx=px(10), pady=(0, px(10)))
        ctk.CTkButton(bot, text="บันทึกทั้งหมดอีกครั้ง", command=self._save_all).pack(side="right", padx=px(6))
        ctk.CTkButton(bot, text="ปิด", command=self.destroy).pack(side="right")

        self.current_path = None
        self.headers = []
        self.header_row_idx = 1
        self.weekday_names = {"วันจันทร์", "วันอังคาร", "วันพุธ", "วันพฤหัสบดี", "วันศุกร์", "วันเสาร์", "วันอาทิตย์"}

        self._refresh_file_list()

    def _refresh_file_list(self):
        prev = self.combo.get()
        paths = sorted(
            glob.glob(os.path.join(self.folder, "Weekly_*.xlsx")),
            key=lambda p: os.path.getmtime(p),
            reverse=True,
        )
        names = [os.path.basename(p) for p in paths]
        self.combo["values"] = names
        if not names:
            self.combo.set("")
            self.current_path = None
            self.tree.delete(*self.tree.get_children())
            self.title("Excel Viewer - Weekly (ไม่มีไฟล์)")
            return
        if prev in names:
            self.combo.current(names.index(prev))
        else:
            self.combo.current(0)
        self._load_selected()

    def _load_selected(self):
        name = self.combo.get()
        if not name:
            return
        path = os.path.join(self.folder, name)
        if not os.path.exists(path):
            messagebox.showerror("Excel", f"ไม่พบไฟล์: {path}")
            return
        self.current_path = path
        self._load_excel_to_tree(path)

    def _detect_header_row(self, sh):
        for r in range(1, min(sh.max_row, 30) + 1):
            v = sh.cell(row=r, column=1).value
            if isinstance(v, str) and v.strip() == "วันที่":
                return r
        return 1

    def _load_excel_to_tree(self, path):
        wb = openpyxl.load_workbook(path, data_only=True)
        sh = wb.active

        self.header_row_idx = self._detect_header_row(sh)
        self.headers = [cell.value if cell.value is not None else "" for cell in sh[self.header_row_idx]]
        if not self.headers:
            self.headers = ["วันที่", "รหัสชุด", "จำนวนจานทั้งหมด", "ไม่มีตำหนิ", "มีตำหนิ", "ตำหนิที่พบบ่อยที่สุด", "หมายเหตุ"]

        self.tree.delete(*self.tree.get_children())
        self.protected_rows.clear()
        self.tree["columns"] = [str(i) for i in range(len(self.headers))]
        for i, h in enumerate(self.headers):
            self.tree.heading(str(i), text=str(h))
            self.tree.column(str(i), width=150 if i == 0 else 130, anchor="center")

        for r in range(self.header_row_idx + 1, sh.max_row + 1):
            row_vals = [sh.cell(row=r, column=c).value for c in range(1, len(self.headers) + 1)]
            row_vals = [("" if v is None else str(v)) for v in row_vals]
            is_day_header = (row_vals[0] in self.weekday_names) and all(v == "" for v in row_vals[1:])
            if any(v != "" for v in row_vals):
                if is_day_header:
                    iid = self.tree.insert("", "end", values=row_vals, tags=("day_header",))
                    self.protected_rows.add(iid)
                else:
                    self.tree.insert("", "end", values=row_vals)

        wb.close()
        self.title(f"Excel Viewer - Weekly - {os.path.basename(path)}")

    def _begin_edit(self, event):
        if not self.current_path:
            return
        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if not row_id or not col_id:
            return
        if row_id in self.protected_rows:
            return

        col_idx = int(col_id.replace("#", "")) - 1
        x, y, w, h = self.tree.bbox(row_id, col_id)
        cur = self.tree.set(row_id, col_id)

        entry = tk.Entry(self.tree)
        entry.insert(0, cur)
        entry.select_range(0, tk.END)
        entry.focus()
        entry.place(x=x, y=y, width=w, height=h)

        def commit(_=None):
            new_val = entry.get()
            entry.destroy()
            values = list(self.tree.item(row_id, "values"))
            values[col_idx] = new_val
            self.tree.item(row_id, values=values)
            self._write_back_cell(row_id, col_idx, new_val)

        def cancel(_=None):
            entry.destroy()

        entry.bind("<Return>", commit)
        entry.bind("<Escape>", cancel)
        entry.bind("<FocusOut>", commit)

    def _tree_index_of(self, row_id):
        return self.tree.get_children().index(row_id)

    def _write_back_cell(self, row_id, col_idx, value):
        try:
            wb = openpyxl.load_workbook(self.current_path)
            sh = wb.active
            excel_row = self.header_row_idx + 1 + self._tree_index_of(row_id)
            excel_col = col_idx + 1
            sh.cell(row=excel_row, column=excel_col, value=value)
            wb.save(self.current_path)
            wb.close()
        except Exception as e:
            print(f"[Weekly write-back] {e}")

    def _save_all(self):
        if not self.current_path:
            return
        try:
            wb = openpyxl.load_workbook(self.current_path)
            sh = wb.active
            r0 = self.header_row_idx + 1
            for i, row_id in enumerate(self.tree.get_children(), start=r0):
                if row_id in self.protected_rows:
                    continue
                vals = list(self.tree.item(row_id, "values"))
                for j, v in enumerate(vals, start=1):
                    sh.cell(row=i, column=j, value=v)
            wb.save(self.current_path)
            wb.close()
            messagebox.showinfo("Excel", "บันทึกเรียบร้อย")
        except Exception as e:
            messagebox.showerror("Excel", f"บันทึกไม่สำเร็จ:\n{e}")


# ================================
# Main App
# ================================
class LeafPlateDetectionApp:
    """
    - รองรับสเกลอัตโนมัติให้พอดีกับจอ MacBook Air M1 (และจออื่น)
    - ปุ่ม “รายงาน” เลือกได้: รายวัน / รายสัปดาห์
    - Weekly viewer: มีหัววันสีเทาและล็อกแก้ไข
    """

    # ---------------- Fonts ----------------
    def setup_fonts(self):
        def scaled(size):
            return max(10, int(round(size * self.SCALE)))

        self.FONT_FAMILY = "Arial"
        for name in (
            "TkDefaultFont", "TkHeadingFont", "TkTextFont", "TkMenuFont",
            "TkFixedFont", "TkTooltipFont", "TkCaptionFont",
            "TkSmallCaptionFont", "TkIconFont"
        ):
            try:
                tkfont.nametofont(name).configure(
                    family=self.FONT_FAMILY,
                    size=scaled(tkfont.nametofont(name).cget("size"))
                )
            except tk.TclError:
                pass

        self.F = lambda size, bold=False: ctk.CTkFont(
            family=self.FONT_FAMILY, size=scaled(size), weight=("bold" if bold else "normal")
        )
        self.FTK = lambda size, bold=False: (
            (self.FONT_FAMILY, scaled(size), "bold") if bold else (self.FONT_FAMILY, scaled(size))
        )

    def _emoji_font(self, size=36):
        size = max(12, int(round(size * self.SCALE)))
        if sys.platform.startswith("win"):
            for font_name in ["Segoe UI Emoji", "Microsoft YaHei UI", "Segoe UI Symbol", "Apple Color Emoji"]:
                try:
                    return ctk.CTkFont(family=font_name, size=size, weight="bold")
                except:
                    continue
            return ctk.CTkFont(size=size, weight="bold")
        elif sys.platform == "darwin":
            return ctk.CTkFont(family="Apple Color Emoji", size=size, weight="bold")
        else:
            for font_name in ["Noto Color Emoji", "Apple Color Emoji", "EmojiOne Color", "Twemoji"]:
                try:
                    return ctk.CTkFont(family=font_name, size=size, weight="bold")
                except:
                    continue
            return ctk.CTkFont(size=size, weight="bold")

    def _get_colored_emoji(self, emoji_code, color_code=None):
        if EMOJI_AVAILABLE:
            try:
                return emojis.encode(emoji_code)
            except:
                pass
        emoji_map = {
            ":red_circle:": "🔴",
            ":green_square:": "🟩",
            ":blue_circle:": "🔵",
            ":white_check_mark:": "✅",
            ":cross_mark:": "❌",
            ":warning:": "⚠️",
            ":information_source:": "ℹ️",
        }
        return emoji_map.get(emoji_code, emoji_code)

    def _print_colored_emoji_message(self, message, emoji_code, color="white"):
        if COLORAMA_AVAILABLE:
            emoji_char = self._get_colored_emoji(emoji_code)
            color_func = getattr(Fore, color.upper(), Fore.WHITE)
            print(f"{color_func}{emoji_char} {message}{Style.RESET_ALL}")
        else:
            emoji_char = self._get_colored_emoji(emoji_code)
            print(f"{emoji_char} {message}")

    def _log_with_emoji(self, level, message):
        emoji_map = {
            "info": ":information_source:",
            "success": ":white_check_mark:",
            "warning": ":warning:",
            "error": ":cross_mark:",
            "debug": ":information_source:"
        }
        color_map = {"info": "cyan", "success": "green", "warning": "yellow", "error": "red", "debug": "blue"}
        self._print_colored_emoji_message(message, emoji_map.get(level, ":information_source:"), color_map.get(level, "white"))

    # ---------------- Data / Config ----------------
    def initialize_data(self):
        # ==== สเกลอัตโนมัติ ====
        self.BASE_W, self.BASE_H = 1920, 1080
        _tmp = tk.Tk()
        try:
            screen_w = _tmp.winfo_screenwidth()
            screen_h = _tmp.winfo_screenheight()
        finally:
            _tmp.destroy()

        margin_w = 40
        margin_h = 80
        usable_w = max(1024, screen_w - margin_w)
        usable_h = max(700, screen_h - margin_h)

        self.SCALE = min(usable_w / self.BASE_W, usable_h / self.BASE_H)
        def px(v): return int(round(v * self.SCALE))
        self.px = px

        # Layout
        self.W, self.H = px(self.BASE_W), px(self.BASE_H)
        self.M = px(25)
        self.HEADER_Y, self.HEADER_H = px(20), px(90)
        self.TOP_Y, self.TOP_H = px(120), px(640)
        self.GAP = self.M
        self.header_w = self.W - 2 * self.M
        self.left_w = int(self.header_w * 0.58)
        self.right_w = self.header_w - self.left_w - self.GAP
        self.BOTTOM_Y = self.TOP_Y + self.TOP_H + px(20)
        self.BOTTOM_H = self.H - self.BOTTOM_Y - self.M
        self.bottom_w = self.header_w
        self.cam_pad = px(20)
        self.cam_h = px(540)
        self.cam_w = self.left_w - (self.cam_pad * 2)

        # ปุ่มสี (พาเลต)
        self.COLOR_PRIMARY = "#2563EB"         # Blue-600
        self.COLOR_PRIMARY_HOVER = "#1E40AF"   # Blue-800
        self.COLOR_DANGER = "#DC2626"          # Red-600
        self.COLOR_DANGER_HOVER = "#B91C1C"    # Red-700
        self.COLOR_NEUTRAL = "#6B7280"         # Gray-500
        self.COLOR_NEUTRAL_HOVER = "#4B5563"   # Gray-600
        self.COLOR_ACCENT = "#F59E0B"          # Amber-500
        self.COLOR_ACCENT_HOVER = "#D97706"    # Amber-600

        # Counters/labels
        self.shape_counts = {"heart": 0, "rectangle": 0, "circle": 0, "total": 0}
        self.defect_data = [
            ("รอยแตก", "ยังไม่พบ", "green"),
            ("รู", "ยังไม่พบ", "green"),
            ("รอยขีดข่วน", "ยังไม่พบ", "green"),
            ("รอยไหม้", "ยังไม่พบ", "green"),
        ]
        self.status_labels = {}
        self.defect_th_map = {"crack": "รอยแตก", "hole": "รู", "bulge": "รอยขีดข่วน", "burn": "รอยไหม้"}
        self._defect_defaults = {d: (s, c) for d, s, c in self.defect_data}

        self.is_collecting_data = False
        self.camera_running = False
        self.cap = None

        self.session_rows = []
        self.session_meta = {}
        self.plate_id_counter = 1
        
        # ต้องกำหนด BASE_DIR และ save_root ก่อนเรียก generate_lot_id()
        self.BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        self.save_root = os.path.join(self.BASE_DIR, "savefile")
        os.makedirs(self.save_root, exist_ok=True)
        
        # ตอนนี้เรียก generate_lot_id() ได้แล้ว
        self.lot_id = self.generate_lot_id()

        # YOLO (separate models for shape and defect)
        self.SHAPE_MODEL_PATH = os.path.join(self.BASE_DIR, "shape_best_rf.pt")
        self.DEFECT_MODEL_PATH = os.path.join(self.BASE_DIR, "defect.pt")
        self.shape_model = None
        self.defect_model = None
        self.imgsz = 896
        # Separate thresholds for better stability
        self.shape_conf_thr = 0.58  # stricter for shapes to avoid false class overlap
        self.defect_conf_thr = 0.27
        self.iou_thr = 0.65
        self.shape_nms_iou = 0.60  # NMS IoU for shapes

        # class groups
        self.shape_classes = {"circle_leaf_plate", "heart_shaped_leaf_plate", "rectangular_leaf_plate"}
        self.defect_classes = {"crack", "hole", "bulge", "burn"}
        self.shape_map = {
            "heart_shaped_leaf_plate": "heart",
            "rectangular_leaf_plate": "rectangle",
            "circle_leaf_plate": "circle",
        }
        self.shape_display_map = {"heart": "หัวใจ", "rectangle": "สี่เหลี่ยมผืนผ้า", "circle": "วงกลม"}

        # files
        self.captures_dir = os.path.join(self.BASE_DIR, "captures")
        os.makedirs(self.captures_dir, exist_ok=True)
        self._auto_xlsx_path = None
        self._auto_json_path = None
        self._session_stamp = None

        # Firebase
        self.firebase_base = "https://leaf-plate-defect-detec-w-cnn-default-rtdb.asia-southeast1.firebasedatabase.app"
        self.firebase_session_key = None
        self._fb_ready = False

        # gating
        self.lbl_heart = self.lbl_rect = self.lbl_circle = None
        self.gate_has_plate = False
        self.gate_has_counted = False
        self.gate_present_frames = 0
        self.gate_absent_frames = 0
        self.gate_present_thresh = 1
        self.gate_absent_thresh = 2
        self.lbl_plate_status = None
        self._latched_defect_counts = {"crack": 0, "hole": 0, "bulge": 0, "burn": 0}
        self._latched_shapes = set()
        self._freeze_after_count = False  # when true, ignore new detections until plate leaves
        self._last_shape_bbox = None      # (x1,y1,x2,y2) of best shape
        self._shape_change_frames = 0
        self.shape_change_thresh = 5
        self.shape_change_iou = 0.15
        self._shape_move_frames = 0
        # require a noticeable movement for several frames to consider new plate
        self.shape_move_frames_thresh = 3
        self.shape_move_dist_px = max(20, int(min(self.cam_w, self.cam_h) * 0.12))
        # ROI texture-change based new-plate detection
        self._prev_shape_roi_gray = None
        self._roi_change_frames = 0
        self.roi_change_frames_thresh = 3
        self.roi_change_mae_thresh = 10.0  # mean abs diff 0-255
        # shape voting per plate
        self._shape_votes = Counter()
        
        # เพิ่มตัวแปรสำหรับการติดตามสถานะแบบไดนามิก
        self._plate_final_status = None  # None, "pass", "defect"
        self._defect_detected_flag = False  # ถ้าเป็น True แล้วจะไม่เปลี่ยนกลับเป็น False
        self._current_defect_count = 0
        self._plate_counted_already = False  # ติดตามว่านับจานไปแล้วหรือยัง

    def generate_lot_id(self):
        """สร้าง lot_id โดยตรวจสอบข้อมูลเดิมในวันเดียวกันและนับต่อจากชุดล่าสุด"""
        today_str = datetime.now().strftime("%y%m%d")
        base = f"PTP{today_str}"
        
        # ตรวจสอบไฟล์ Excel ที่มีอยู่ในวันเดียวกัน
        existing_lots = self._get_existing_lots_today()
        
        if existing_lots:
            # หาหมายเลขชุดล่าสุดและนับต่อ
            max_seq = max(existing_lots)
            new_seq = max_seq + 1
        else:
            # ไม่มีข้อมูลเดิม เริ่มต้นจากชุดที่ 1
            new_seq = 1
        
        return f"{base}_{new_seq:02d}"

    def _get_existing_lots_today(self):
        """ตรวจสอบไฟล์ Excel ที่มีอยู่ในวันนี้และคืนค่าลิสต์ของหมายเลขชุดที่มีอยู่"""
        # เพิ่มการตรวจสอบความปลอดภัย
        if not hasattr(self, 'save_root') or not self.save_root:
            return []
            
        today_str = datetime.now().strftime("%y%m%d")
        today_date = self.thai_date(datetime.now())
        base_pattern = f"PTP{today_str}"
        existing_lots = set()
        
        # ตรวจสอบไฟล์ Report_*.xlsx ทั้งหมดในโฟลเดอร์
        for report_file in glob.glob(os.path.join(self.save_root, "Report_*.xlsx")):
            try:
                wb = openpyxl.load_workbook(report_file, data_only=True)
                sh = wb.active
                
                # หาแถว header
                header_row = 1
                for r in range(1, min(sh.max_row, 20) + 1):
                    val = sh.cell(row=r, column=1).value
                    if isinstance(val, str) and val.strip() == "วันที่":
                        header_row = r
                        break
                
                # อ่านข้อมูลทุกแถว
                for r in range(header_row + 1, sh.max_row + 1):
                    date_cell = sh.cell(row=r, column=1).value
                    lot_cell = sh.cell(row=r, column=4).value  # คอลัมน์รหัสชุด
                    
                    if not date_cell or not lot_cell:
                        continue
                        
                    # ตรวจสอบว่าเป็นวันเดียวกันหรือไม่
                    if str(date_cell).strip() == today_date:
                        lot_str = str(lot_cell).strip()
                        # ตรวจสอบรูปแบบ PTP + วันที่ + _XX
                        if lot_str.startswith(base_pattern) and "_" in lot_str:
                            try:
                                seq_part = lot_str.split("_")[-1]
                                seq_num = int(seq_part)
                                existing_lots.add(seq_num)
                            except (ValueError, IndexError):
                                continue
                
                wb.close()
                
            except Exception as e:
                print(f"Error reading Excel file {report_file}: {e}")
                continue
        
        return list(existing_lots)

    def _check_and_continue_session(self):
        """ตรวจสอบว่ามีเซสชันในวันเดียวกันหรือไม่ และอัปเดตข้อมูลให้เหมาะสม"""
        # เพิ่มการตรวจสอบความปลอดภัย
        if not hasattr(self, 'save_root') or not self.save_root:
            return
            
        today_str = datetime.now().strftime("%Y%m%d")
        today_date = self.thai_date(datetime.now())
        
        # หาไฟล์รายงานของวันนี้
        today_reports = []
        for report_file in glob.glob(os.path.join(self.save_root, f"Report_{today_str}_*.xlsx")):
            today_reports.append(report_file)
        
        if today_reports:
            # เรียงตามเวลาล่าสุด
            today_reports.sort(key=lambda x: os.path.getmtime(x), reverse=True)
            latest_report = today_reports[0]
            
            try:
                wb = openpyxl.load_workbook(latest_report, data_only=True)
                sh = wb.active
                
                # หาแถว header
                header_row = 1
                for r in range(1, min(sh.max_row, 20) + 1):
                    val = sh.cell(row=r, column=1).value
                    if isinstance(val, str) and val.strip() == "วันที่":
                        header_row = r
                        break
                
                # หาจานที่มีหมายเลขสูงสุดในวันนี้
                max_plate_id = 0
                for r in range(header_row + 1, sh.max_row + 1):
                    date_cell = sh.cell(row=r, column=1).value
                    plate_cell = sh.cell(row=r, column=3).value  # คอลัมน์จานที่
                    
                    if not date_cell or not plate_cell:
                        continue
                        
                    if str(date_cell).strip() == today_date:
                        try:
                            plate_num = int(str(plate_cell).strip())
                            max_plate_id = max(max_plate_id, plate_num)
                        except (ValueError, TypeError):
                            continue
                
                wb.close()
                
                # อัปเดต counter ให้นับต่อจากหมายเลขล่าสุด
                if max_plate_id > 0:
                    self.plate_id_counter = max_plate_id + 1
                    print(f"พบข้อมูลเดิมในวันนี้ จะนับต่อจากจานที่ {self.plate_id_counter}")
                    
            except Exception as e:
                print(f"Error checking existing session: {e}")

    # ---------------- Tk error patch ----------------
    def _report_callback_exception(self, exc, val, tb):
        import traceback
        if exc is KeyboardInterrupt or isinstance(val, KeyboardInterrupt):
            print("[INFO] KeyboardInterrupt in Tk callback (ignored). Use the UI buttons to stop.")
            return
        traceback.print_exception(exc, val, tb)
        try:
            messagebox.showerror("Unexpected error", f"{exc.__name__}: {val}")
        except Exception:
            pass

    # ---------------- Firebase helpers ----------------
    def _fb_init(self):
        if self._fb_ready:
            return
        try:
            cred_path = os.path.join(self.BASE_DIR, "serviceAccountKey.json")
            cred = credentials.Certificate(cred_path)
            firebase_admin.initialize_app(cred, {"databaseURL": self.firebase_base})
            self._fb_ready = True
        except Exception as e:
            print(f"[Firebase] Admin init failed, will fallback to REST. reason={e}")
            self._fb_ready = False

    def _firebase_post(self, path, obj):
        payload = {
            **obj,
            "_meta": {"source": "python-admin", "pushed_at": datetime.now(timezone.utc).isoformat(), "server_id": str(uuid.uuid4())}
        }
        try:
            self._fb_init()
            if self._fb_ready:
                db.reference(path).push(payload)
                return
        except Exception as e:
            print(f"[Firebase] Admin push failed, fallback REST. reason={e}")
        try:
            import urllib.request
            url = f"{self.firebase_base}/{path}.json"
            data = json.dumps(payload).encode("utf-8")
            req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"})
            with urllib.request.urlopen(req, timeout=5): pass
        except Exception as e:
            print(f"[Firebase] REST POST error: {e}")

    def _firebase_put(self, path, obj):
        try:
            self._fb_init()
            if self._fb_ready:
                db.reference(path).set(obj)
                return
        except Exception as e:
            print(f"[Firebase] Admin set failed, fallback REST. reason={e}")
        try:
            import urllib.request
            url = f"{self.firebase_base}/{path}.json"
            data = json.dumps(obj).encode("utf-8")
            req = urllib.request.Request(url, data=data, headers={"Content-Type": "application/json"}, method="PUT")
            with urllib.request.urlopen(req, timeout=5): pass
        except Exception as e:
            print(f"[Firebase] REST PUT error: {e}")

    # ---------------- App / Camera / Model ----------------
    def setup_app(self):
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        self.app = ctk.CTk()
        self.app.title("Leaf Plate Defect Detection")

        # macOS: ปิด tk scaling เพื่อไม่ให้ซ้ำกับ Retina แล้วใช้ self.SCALE คุมเอง
        try:
            if sys.platform == "darwin":
                self.app.tk.call('tk', 'scaling', 1.0)
        except Exception:
            pass

        self.app.geometry(f"{self.W}x{self.H}+20+40")
        self.app.resizable(False, False)
        self.app.configure(fg_color="#ffffff")
        self.app.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.app.report_callback_exception = self._report_callback_exception

    def setup_camera(self):
        if sys.platform.startswith("win"):
            backend = cv2.CAP_DSHOW
        elif sys.platform == "darwin":
            backend = cv2.CAP_AVFOUNDATION
        else:
            backend = cv2.CAP_V4L2
        self.cap = cv2.VideoCapture(0, backend)
        if not self.cap or not self.cap.isOpened():
            self.cap = cv2.VideoCapture(0)
        if not self.cap or not self.cap.isOpened():
            print("Cannot open camera"); self.cap = None; return
        target_w = max(640, min(1280, self.cam_w))
        target_h = max(360, min(720, self.cam_h))
        self.cap.set(cv2.CAP_PROP_FRAME_WIDTH, target_w)
        self.cap.set(cv2.CAP_PROP_FRAME_HEIGHT, target_h)

    def setup_model(self):
        shape_ok = defect_ok = False
        try:
            self.shape_model = YOLO(self.SHAPE_MODEL_PATH)
            shape_ok = True
            self._log_with_emoji("success", "โหลดโมเดลรูปทรง (shape.pt) สำเร็จ")
        except Exception as e:
            self._log_with_emoji("error", f"โหลดโมเดลรูปทรงไม่สำเร็จ: {e}")
            self.shape_model = None
        try:
            self.defect_model = YOLO(self.DEFECT_MODEL_PATH)
            defect_ok = True
            self._log_with_emoji("success", "โหลดโมเดลตำหนิ (defect.pt) สำเร็จ")
        except Exception as e:
            self._log_with_emoji("error", f"โหลดโมเดลตำหนิไม่สำเร็จ: {e}")
            self.defect_model = None
        if not (shape_ok and defect_ok):
            messagebox.showerror("Model Error", "โหลดโมเดลไม่ครบ กรุณาตรวจสอบไฟล์ shape.pt และ defect.pt")

    # ---------------- UI ----------------
    def create_widgets(self):
        self.create_header()
        self.create_main_content()

    def create_header(self):
        px = self.px
        header_frame = ctk.CTkFrame(self.app, width=self.header_w, height=self.HEADER_H,
                                    fg_color="#7A5429", corner_radius=px(10))
        header_frame.place(x=self.M, y=self.HEADER_Y)

        ctk.CTkLabel(header_frame, text="โปรแกรมตรวจจับรอยตำหนิบนจานใบไม้",
                     font=self.F(32, True), text_color="white").place(x=px(50), y=px(28))

        self.header_time_label = ctk.CTkLabel(header_frame, text=datetime.now().strftime("%H:%M:%S"),
                                              font=self.F(22, True), text_color="white")
        self.header_time_label.place(x=self.header_w - px(220), y=px(15))

        thai_year = datetime.now().year + 543
        date_str = datetime.now().strftime(f"%d/%m/{thai_year}")
        self.header_date_label = ctk.CTkLabel(header_frame, text=date_str,
                                              font=self.F(18), text_color="white")
        self.header_date_label.place(x=self.header_w - px(220), y=px(50))
        self.update_header_time()

    def create_main_content(self):
        px = self.px
        # left
        left_frame = ctk.CTkFrame(self.app, width=self.left_w, height=self.TOP_H,
                                  fg_color="#ffffff", corner_radius=px(15),
                                  border_width=px(2), border_color="#7A5429")
        left_frame.place(x=self.M, y=self.TOP_Y)

        self.camera_frame = ctk.CTkFrame(left_frame, width=self.cam_w, height=self.cam_h,
                                         fg_color="#E4DFDA", corner_radius=px(10),
                                         border_width=px(1), border_color="#7A5429")
        self.camera_frame.place(x=self.cam_pad, y=self.cam_pad)

        self.camera_label = tk.Label(self.camera_frame, text="Initializing Camera...",
                                     font=self.FTK(16), fg="black", bg="#7A5429")
        self.camera_label.place(x=0, y=0, width=self.cam_w, height=self.cam_h)

        button_frame_w = self.left_w - 2 * self.cam_pad
        button_frame = ctk.CTkFrame(left_frame, width=button_frame_w, height=px(60), fg_color="transparent")
        button_frame.place(x=self.cam_pad, y=self.cam_pad + self.cam_h + px(15))

        self.toggle_button = ctk.CTkButton(
            button_frame, width=px(150), height=px(45), text="เริ่ม",
            font=self.F(16, True), text_color="#FFFFFF",
            fg_color=self.COLOR_PRIMARY, hover_color=self.COLOR_PRIMARY_HOVER,
            command=self.toggle_data_collection
        )
        self.toggle_button.place(x=0, y=px(5))

        # ปุ่ม "รายงาน" -> Popup เลือกรายวัน/รายสัปดาห์
        ctk.CTkButton(
            button_frame, width=px(170), height=px(45), text="รายงาน",
            font=self.F(16, True), text_color="#FFFFFF",
            fg_color=self.COLOR_PRIMARY, hover_color=self.COLOR_PRIMARY_HOVER,
            command=self.show_report_picker
        ).place(x=button_frame_w - px(170), y=px(5))

        # right
        right_x = self.M + self.left_w + self.GAP
        right_frame = ctk.CTkFrame(self.app, width=self.right_w, height=self.TOP_H,
                                   fg_color="#E4DFDA", corner_radius=px(15),
                                   border_width=px(2), border_color="#7A5429")
        right_frame.place(x=right_x, y=self.TOP_Y)
        self.create_unified_panel(right_frame)

        # bottom
        bottom_frame = ctk.CTkFrame(self.app, width=self.bottom_w, height=self.BOTTOM_H,
                                    fg_color="#ffffff", corner_radius=px(15),
                                    border_width=px(2), border_color="#7A5429")
        bottom_frame.place(x=self.M, y=self.BOTTOM_Y)
        self.create_defect_table(bottom_frame)

    # ----- popup เลือกประเภทรายงาน -----
    def show_report_picker(self):
        px = self.px
        dlg = ctk.CTkToplevel(self.app)
        dlg.title("เลือกรูปแบบรายงาน")
        dlg.geometry(f"{px(420)}x{px(210)}")
        dlg.resizable(False, False)
        dlg.transient(self.app)
        dlg.grab_set()

        ctk.CTkLabel(dlg, text="ต้องการดูรายงานแบบใด?", font=self.F(20, True)).pack(pady=(px(18), px(10)))

        row = ctk.CTkFrame(dlg, fg_color="transparent")
        row.pack(pady=px(8))

        # รายงาน (รายวัน)
        ctk.CTkButton(
            row, width=px(160), height=px(48), text="รายงาน (รายวัน)",
            font=self.F(16, True), text_color="#FFFFFF",
            fg_color=self.COLOR_PRIMARY, hover_color=self.COLOR_PRIMARY_HOVER,
            command=lambda: (dlg.destroy(), self.open_excel_viewer())
        ).pack(side="left", padx=px(8))

        # รายงาน (รายสัปดาห์)
        ctk.CTkButton(
            row, width=px(160), height=px(48), text="รายงาน (รายสัปดาห์)",
            font=self.F(16, True), text_color="#FFFFFF",
            fg_color=self.COLOR_ACCENT, hover_color=self.COLOR_ACCENT_HOVER,
            command=lambda: (dlg.destroy(), self.open_weekly_viewer())
        ).pack(side="left", padx=px(8))

        ctk.CTkButton(
            dlg, text="ปิด", width=px(100),
            font=self.F(14, True), text_color="#FFFFFF",
            fg_color=self.COLOR_NEUTRAL, hover_color=self.COLOR_NEUTRAL_HOVER,
            command=dlg.destroy
        ).pack(pady=(px(10), px(16)))

    def metric_box(self, panel, x, box_w, box_h, box_y, emoji, emoji_font, title_text, title_color, count_label=None):
        px = self.px
        f = ctk.CTkFrame(panel, width=box_w, height=box_h, fg_color="#ffffff", corner_radius=px(18))
        f.place(x=x, y=box_y)
        ctk.CTkLabel(f, text=emoji, font=emoji_font, text_color=title_color)\
            .place(x=box_w // 2, y=px(28), anchor="center")
        if count_label:
            count_label.place(x=box_w // 2, y=px(90), anchor="center")
        return f

    def create_unified_panel(self, parent):
        px = self.px
        panel_w = self.right_w - px(40)
        panel_h = self.TOP_H - px(40)
        panel = ctk.CTkFrame(parent, width=panel_w, height=panel_h, fg_color="#E4DFDA", corner_radius=px(15))
        panel.place(x=px(20), y=px(20))

        ctk.CTkLabel(panel, text="บันทึกได้ทั้งหมด", font=self.F(26, True), text_color="#1a2a3a")\
            .place(x=panel_w // 2, y=px(40), anchor="center")

        self.total_number_label = ctk.CTkLabel(panel, text=str(self.shape_counts["total"]),
                                               font=self.F(70, True), text_color="#e74c3c")
        self.total_number_label.place(x=panel_w // 2, y=px(110), anchor="center")

        ctk.CTkLabel(panel, text="จานแต่ละรูปทรง", font=self.F(20, True), text_color="#1a2a3a")\
            .place(x=panel_w // 2, y=px(180), anchor="center")

        # cards
        box_w, box_h, box_y = px(220), px(120), px(220)
        gap = (panel_w - 3 * box_w) // 4
        emoji_font = self._emoji_font(36)

        f1 = self.metric_box(panel, gap, box_w, box_h, box_y, "♥", emoji_font, "หัวใจ", "#e74c3c")
        self.lbl_heart = ctk.CTkLabel(f1, text="0", font=self.F(40, True), text_color="#102438")
        self.lbl_heart.place(x=box_w // 2, y=px(90), anchor="center")

        f2 = self.metric_box(panel, gap * 2 + box_w, box_w, box_h, box_y, "▬", emoji_font, "สี่เหลี่ยมผืนผ้า", "#199129")
        self.lbl_rect = ctk.CTkLabel(f2, text="0", font=self.F(40, True), text_color="#102438")
        self.lbl_rect.place(x=box_w // 2, y=px(90), anchor="center")

        f3 = self.metric_box(panel, gap * 3 + box_w * 2, box_w, box_h, box_y, "●", emoji_font, "วงกลม", "#2AA7B8")
        self.lbl_circle = ctk.CTkLabel(f3, text="0", font=self.F(40, True), text_color="#102438")
        self.lbl_circle.place(x=box_w // 2, y=px(90), anchor="center")

        # summary card
        summary_card_y = box_y + box_h + px(30)
        summary_card_h = panel_h - summary_card_y - px(20)
        summary_card_w = panel_w - 2 * gap
        summary = ctk.CTkFrame(panel, width=summary_card_w, height=summary_card_h, fg_color="#ffffff", corner_radius=px(18))
        summary.place(x=gap, y=summary_card_y)

        ctk.CTkLabel(summary, text="สรุปผลการตรวจ", font=self.F(20, True), text_color="#1a2a3a")\
            .place(x=summary_card_w // 2, y=px(32), anchor="center")

        label_x = px(60)
        value_x = summary_card_w // 2 + px(20)
        start_y = px(70)
        row_h = px(38)

        self.lbl_plate_order = ctk.CTkLabel(summary, text="0", font=self.F(18, True), text_color="#1a2a3a")
        self.lbl_plate_order.place(x=value_x, y=start_y, anchor="w")
        ctk.CTkLabel(summary, text="ลำดับจาน:", font=self.F(16), text_color="#1a2a3a")\
            .place(x=label_x, y=start_y, anchor="w")

        self.lbl_plate_status = ctk.CTkLabel(summary, text="ยังไม่ได้ตรวจ", font=self.F(18, True), text_color="#888888")
        self.lbl_plate_status.place(x=value_x, y=start_y + row_h, anchor="w")
        ctk.CTkLabel(summary, text="สถานะ:", font=self.F(16), text_color="#1a2a3a")\
            .place(x=label_x, y=start_y + row_h, anchor="w")

        self.lbl_defect_count = ctk.CTkLabel(summary, text="ยังไม่พบ", font=self.F(20, True), text_color="#888888")
        self.lbl_defect_count.place(x=value_x, y=start_y + row_h * 2, anchor="w")
        ctk.CTkLabel(summary, text="จำนวนข้อบกพร่อง:", font=self.F(16), text_color="#1a2a3a")\
            .place(x=label_x, y=start_y + row_h * 2, anchor="w")

        self.lbl_lot = ctk.CTkLabel(summary, text=f"{self.lot_id}", font=self.F(16, True), text_color="#1a2a3a")
        self.lbl_lot.place(x=value_x, y=start_y + row_h * 3, anchor="w")
        ctk.CTkLabel(summary, text="รหัสชุด:", font=self.F(16), text_color="#1a2a3a")\
            .place(x=label_x, y=start_y + row_h * 3, anchor="w")

    def create_defect_table(self, parent):
        px = self.px
        table_w = self.bottom_w - px(40)
        table_h = self.BOTTOM_H - px(40)
        table_frame = ctk.CTkFrame(parent, width=table_w, height=table_h, fg_color="#ffffff", corner_radius=px(10))
        table_frame.place(x=px(20), y=px(20))

        header = ctk.CTkFrame(table_frame, width=table_w - px(10), height=px(48),
                              fg_color="#7A5429", corner_radius=px(8))
        header.place(x=px(5), y=px(5))

        col1_x = int((table_w - px(10)) * 0.32)
        col2_x = int((table_w - px(10)) * 0.76)

        ctk.CTkLabel(header, text="ประเภทตำหนิ", font=self.F(18, True), text_color="white")\
            .place(x=col1_x, y=px(24), anchor="center")
        ctk.CTkLabel(header, text="สถานะตำหนิ", font=self.F(18, True), text_color="white")\
            .place(x=col2_x, y=px(24), anchor="center")

        row_start_y, row_h = px(58), px(42)
        for i, (defect, status, color) in enumerate(self.defect_data):
            y = row_start_y + i * row_h
            row_color = "#ffffff" if i % 2 == 0 else "#EDEAE6"
            rf = ctk.CTkFrame(table_frame, width=table_w - px(10), height=row_h, fg_color=row_color, corner_radius=0)
            rf.place(x=px(5), y=y)
            ctk.CTkLabel(rf, text=defect, font=self.F(17, True), text_color="#1a2a3a")\
                .place(x=col1_x, y=row_h // 2, anchor="center")

            status_color = "#199129" if color == "green" else "#e74c3c"
            lbl = ctk.CTkLabel(rf, text=status, font=self.F(17, True), text_color=status_color)
            lbl.place(x=col2_x, y=row_h // 2, anchor="center")
            self.status_labels[defect] = lbl

    # ---------------- Excel autosave + JSON ----------------
    EXCEL_HEADERS = ["วันที่", "เวลา", "จานที่", "เลขชุด", "รูปทรงจาน", "ตำหนิที่พบ", "หมายเหตุ"]

    def _ensure_session_files(self):
        if self._auto_xlsx_path and self._auto_json_path:
            return
        self._session_stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # ตรวจสอบว่ามีไฟล์รายงานของวันนี้หรือไม่
        today_str = datetime.now().strftime("%Y%m%d")
        existing_today_reports = glob.glob(os.path.join(self.save_root, f"Report_{today_str}_*.xlsx"))
        
        if existing_today_reports:
            # มีไฟล์รายงานวันนี้อยู่แล้ว ใช้ไฟล์เดิม
            existing_today_reports.sort(key=lambda x: os.path.getmtime(x), reverse=True)
            self._auto_xlsx_path = existing_today_reports[0]
            # สร้างไฟล์ JSON ใหม่สำหรับเซสชันนี้
            self._auto_json_path = os.path.join(self.save_root, f"Report_{self._session_stamp}.json")
            
            # อัปเดตเวลาในไฟล์ Excel เดิม
            self._update_excel_session_times()
            self._write_json_to_path(self._auto_json_path)
            
            print(f"ใช้ไฟล์รายงานเดิม: {os.path.basename(self._auto_xlsx_path)}")
        else:
            # ไม่มีไฟล์รายงานวันนี้ สร้างใหม่
            self._auto_xlsx_path = os.path.join(self.save_root, f"Report_{self._session_stamp}.xlsx")
            self._auto_json_path = os.path.join(self.save_root, f"Report_{self._session_stamp}.json")

            wb = openpyxl.Workbook()
            sh = wb.active
            sh.title = "Report"

            title = f"รายงานการตรวจจานใบไม้ วันที่ {self._title_date(datetime.now())}"
            sh.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.EXCEL_HEADERS))
            sh["A1"] = title
            sh["A1"].font = Font(bold=True, size=13)
            sh["A1"].alignment = Alignment(horizontal="left", vertical="center")

            self.session_meta.setdefault("start_time", datetime.now().strftime("%H:%M:%S"))
            sh["A2"] = "เริ่มตรวจ:"; sh["A2"].font = Font(bold=True)
            sh["B2"] = self.session_meta["start_time"]
            sh["D2"] = "สิ้นสุด:";  sh["D2"].font = Font(bold=True)
            sh["E2"] = "-"

            head_row = 4
            sh.append([])
            sh.append(self.EXCEL_HEADERS)

            head_font = Font(bold=True)
            fill = PatternFill("solid", fgColor="E4DFDA")
            align = Alignment(horizontal="center", vertical="center")
            thin = Side(style="thin", color="B7B7B7")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for c in range(1, len(self.EXCEL_HEADERS) + 1):
                cell = sh.cell(row=head_row, column=c)
                cell.font = head_font; cell.fill = fill; cell.alignment = align; cell.border = border
                sh.column_dimensions[get_column_letter(c)].width = 22
            sh.freeze_panes = "A5"

            wb.save(self._auto_xlsx_path); wb.close()
            self._write_json_to_path(self._auto_json_path)
            
            print(f"สร้างไฟล์รายงานใหม่: {os.path.basename(self._auto_xlsx_path)}")
        
        self.firebase_session_key = self._session_stamp

    def _update_excel_session_times(self):
        try:
            wb = openpyxl.load_workbook(self._auto_xlsx_path)
            sh = wb.active
            sh["B2"] = self.session_meta.get("start_time", "")
            sh["E2"] = datetime.now().strftime("%H:%M:%S")
            wb.save(self._auto_xlsx_path)
            wb.close()
        except Exception as e:
            print(f"Update Excel times error: {e}")

    def _append_excel_row(self, row_dict):
        wb = openpyxl.load_workbook(self._auto_xlsx_path)
        sh = wb.active
        vals = [
            row_dict.get("date", ""),
            row_dict.get("time", ""),
            row_dict.get("plate_id", ""),
            row_dict.get("lot_id", ""),
            row_dict.get("shape", "-"),
            row_dict.get("defects", "-"),
            row_dict.get("note", ""),
        ]
        sh.append(vals)
        wb.save(self._auto_xlsx_path)
        wb.close()
        self._update_excel_session_times()

    def _write_json_to_path(self, path):
        payload = {
            "report_title": f"รายงานการตรวจจานใบไม้ วันที่ {self._title_date(datetime.now())}",
            "lot_id": self.lot_id,
            "session": {
                "start_time": self.session_meta.get("start_time"),
                "end_time": datetime.now().strftime("%H:%M:%S")
            },
            "records": self.session_rows
        }
        try:
            with open(path, "w", encoding="utf-8") as jf:
                json.dump(payload, jf, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Write JSON error: {e}")

    # ---------------- Helpers ----------------
    def thai_date(self, dt):  # 11/09/2568
        return dt.strftime(f"%d/%m/{dt.year + 543}")

    def _title_date(self, dt):  # 11/09/68
        return dt.strftime("%d/%m/%y")

    def _update_lot_label(self):
        try:
            self.lbl_lot.configure(text=f"{self.lot_id}")
        except Exception:
            pass

    def _increment_lot_id(self):
        try:
            now_short = datetime.now().strftime("%y%m%d")
            base = f"PTP{now_short}"
            seq = 1
            if isinstance(self.lot_id, str) and "_" in self.lot_id and self.lot_id.startswith("PTP"):
                old_base, old_seq = self.lot_id.split("_", 1)
                if old_base.endswith(now_short):
                    seq = int(old_seq) + 1
            self.lot_id = f"{base}_{seq:02d}"
        except Exception:
            self.lot_id = self.generate_lot_id()
        self._update_lot_label()

    def _reset_defect_table(self):
        for k in list(self._latched_defect_counts.keys()):
            self._latched_defect_counts[k] = 0
        for defect, (status, color) in self._defect_defaults.items():
            lbl = self.status_labels.get(defect)
            if lbl:
                status_color = "#199129" if color == "green" else "#e74c3c"
                lbl.configure(text=status, text_color=status_color)

    def _render_latched_defect_counts(self):
        for en_name, th_name in self.defect_th_map.items():
            lbl = self.status_labels.get(th_name)
            if not lbl: continue
            cnt = int(self._latched_defect_counts.get(en_name, 0))
            if cnt > 0:
                lbl.configure(text=str(cnt), text_color="#e74c3c")
            else:
                default_status, default_color = self._defect_defaults.get(th_name, ("ยังไม่พบ", "green"))
                status_color = "#199129" if default_color == "green" else "#e74c3c"
                lbl.configure(text=default_status, text_color=status_color)

    def _set_plate_status(self, mode, defect_count=None):
        if not self.lbl_plate_status: return
        if mode == "pending":
            self.lbl_plate_status.configure(text="ยังไม่ได้ตรวจ", text_color="#888888")
            self.lbl_defect_count.configure(text="ยังไม่พบ", text_color="#888888")
        elif mode == "counted":
            if defect_count and defect_count > 0:
                self.lbl_plate_status.configure(text="มีตำหนิ", text_color="#e74c3c")
                self.lbl_defect_count.configure(text=str(defect_count), text_color="#e74c3c")
            else:
                self.lbl_plate_status.configure(text="ผ่าน", text_color="#199129")
                self.lbl_defect_count.configure(text="ไม่มีตำหนิ", text_color="#888888")

    # ---------------- Save & Firebase ----------------
    def _save_detection_record(self, annotated_bgr, defect_names: set, shapes_found: set):
        now = datetime.now()
        ts = now.strftime("%Y%m%d_%H%M%S_%f")[:-3]

        img_path = os.path.join(self.captures_dir, f"detect_{ts}.jpg")
        try:
            cv2.imwrite(img_path, annotated_bgr)
        except Exception as e:
            print(f"Save image error: {e}")

        defects_th = [self.defect_th_map.get(d, d) for d in sorted(defect_names)]
        defects_text = "-" if not defects_th else " / ".join(defects_th)

        if not shapes_found:
            shape_text = "-"
        else:
            thai_shapes = [self.shape_display_map.get(s, s) for s in sorted(shapes_found)]
            shape_text = " / ".join(thai_shapes)

        row = {
            "date": self.thai_date(now),
            "time": now.strftime("%H:%M:%S"),
            "plate_id": self.plate_id_counter,
            "lot_id": self.lot_id,
            "shape": shape_text,
            "defects": defects_text.strip(),
            "note": ""
        }
        self.plate_id_counter += 1
        self.session_rows.append(row)

        self._ensure_session_files()
        self._append_excel_row(row)
        self._write_json_to_path(self._auto_json_path)

        meta = {
            "report_title": f"รายงานการตรวจจานใบไม้ วันที่ {self._title_date(datetime.now())}",
            "lot_id": self.lot_id,
            "session": {
                "start_time": self.session_meta.get("start_time"),
                "end_time": datetime.now().strftime("%H:%M:%S")
            }
        }
        self._firebase_put(f"sessions/{self.firebase_session_key}/meta", meta)
        self._firebase_post(f"sessions/{self.firebase_session_key}/records", row)

        return row

    # ---------------- Detection ----------------
    def _annotate_and_summarize(self, frame_bgr, res_shape=None, res_defect=None):
        annotated = frame_bgr.copy()
        shapes_found, defect_names = set(), set()
        defect_counts = {}
        best_shape_bbox = None
        best_shape_conf = -1.0

        # Draw shapes (blue boxes)
        if res_shape is not None and hasattr(res_shape, "boxes") and res_shape.boxes is not None and self.shape_model is not None:
            names_s = self.shape_model.names
            boxes = res_shape.boxes
            xyxy = boxes.xyxy.cpu().numpy()
            clss = boxes.cls.cpu().numpy().astype(int)
            conf = boxes.conf.cpu().numpy()

            # Collect detections and apply simple class-agnostic NMS to avoid overlapping multi-class boxes
            dets = []  # (x1,y1,x2,y2,conf,label)
            for (x1, y1, x2, y2), c, p in zip(xyxy, clss, conf):
                label = names_s.get(int(c), str(c))
                if label not in self.shape_map:
                    continue
                dets.append((float(x1), float(y1), float(x2), float(y2), float(p), label))

            dets.sort(key=lambda d: d[4], reverse=True)
            kept = []
            def iou(a, b):
                ax1, ay1, ax2, ay2 = a[0], a[1], a[2], a[3]
                bx1, by1, bx2, by2 = b[0], b[1], b[2], b[3]
                inter_x1 = max(ax1, bx1)
                inter_y1 = max(ay1, by1)
                inter_x2 = min(ax2, bx2)
                inter_y2 = min(ay2, by2)
                iw = max(0.0, inter_x2 - inter_x1)
                ih = max(0.0, inter_y2 - inter_y1)
                inter = iw * ih
                area_a = max(0.0, (ax2 - ax1)) * max(0.0, (ay2 - ay1))
                area_b = max(0.0, (bx2 - bx1)) * max(0.0, (by2 - by1))
                union = area_a + area_b - inter
                return (inter / union) if union > 0 else 0.0
            for d in dets:
                suppressed = False
                for k in kept:
                    if iou(d, k) >= self.shape_nms_iou:
                        suppressed = True
                        break
                if not suppressed:
                    kept.append(d)

            for x1, y1, x2, y2, p, label in kept:
                xi1, yi1, xi2, yi2 = int(x1), int(y1), int(x2), int(y2)
                cv2.rectangle(annotated, (xi1, yi1), (xi2, yi2), (0, 102, 255), 2)
                cv2.putText(annotated, f"{label} {p:.2f}", (xi1, max(20, yi1 - 6)),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 102, 255), 2, cv2.LINE_AA)
                shapes_found.add(self.shape_map[label])
                if p > best_shape_conf:
                    best_shape_conf = p
                    best_shape_bbox = (xi1, yi1, xi2, yi2)

        # Draw defects (red boxes)
        if res_defect is not None and hasattr(res_defect, "boxes") and res_defect.boxes is not None and self.defect_model is not None:
            names_d = self.defect_model.names
            boxes = res_defect.boxes
            xyxy = boxes.xyxy.cpu().numpy().astype(int)
            clss = boxes.cls.cpu().numpy().astype(int)
            conf = boxes.conf.cpu().numpy()
            for (x1, y1, x2, y2), c, p in zip(xyxy, clss, conf):
                label = names_d.get(int(c), str(c))
                cv2.rectangle(annotated, (x1, y1), (x2, y2), (255, 0, 0), 2)
                cv2.putText(annotated, f"{label} {p:.2f}", (x1, max(20, y1 - 6)),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (20, 20, 255), 2, cv2.LINE_AA)
                if label in self.defect_classes:
                    defect_names.add(label)
                    defect_counts[label] = defect_counts.get(label, 0) + 1

        return annotated, shapes_found, defect_counts, defect_names, best_shape_bbox

    # ---------------- Camera loop ----------------
    def start_camera(self):
        if self.cap:
            self.camera_running = True
            self.update_camera()

    def stop_camera(self):
        self.camera_running = False
        if self.cap:
            self.cap.release(); self.cap = None

    def update_camera(self):
        if not self.camera_running or not self.cap:
            return
        ret, frame = self.cap.read()
        if not ret:
            self.app.after(30, self.update_camera); return

        frame_resized = cv2.resize(frame, (self.cam_w, self.cam_h))

        if self.is_collecting_data and (self.shape_model is not None or self.defect_model is not None):
            try:
                res_shape = None
                res_defect = None
                if self.shape_model is not None:
                    results_s = self.shape_model.predict(
                        source=frame_resized, imgsz=self.imgsz,
                        conf=self.shape_conf_thr, iou=self.iou_thr, verbose=False
                    )
                    res_shape = results_s[0]
                if self.defect_model is not None:
                    results_d = self.defect_model.predict(
                        source=frame_resized, imgsz=self.imgsz,
                        conf=self.defect_conf_thr, iou=self.iou_thr, verbose=False
                    )
                    res_defect = results_d[0]
                annotated, shapes_found, defect_counts, defect_names, best_bbox = self._annotate_and_summarize(frame_resized, res_shape, res_defect)
                frame_to_show = annotated

                # ปรับเงื่อนไขการตรวจจับจานให้รวมทั้งรูปทรงและตำหนิ
                plate_detected = len(shapes_found) > 0 or len(defect_names) > 0
                
                # Detect new plate arrival by low IoU and noticeable movement of best shape bbox
                if plate_detected and best_bbox is not None:
                    def iou_xyxy(a, b):
                        ax1, ay1, ax2, ay2 = a
                        bx1, by1, bx2, by2 = b
                        inter_x1 = max(ax1, bx1); inter_y1 = max(ay1, by1)
                        inter_x2 = min(ax2, bx2); inter_y2 = min(ay2, by2)
                        iw = max(0, inter_x2 - inter_x1); ih = max(0, inter_y2 - inter_y1)
                        inter = iw * ih
                        area_a = max(0, ax2 - ax1) * max(0, ay2 - ay1)
                        area_b = max(0, bx2 - bx1) * max(0, by2 - by1)
                        union = area_a + area_b - inter
                        return (inter / union) if union > 0 else 0.0
                    if self._last_shape_bbox is not None:
                        iou_now = iou_xyxy(best_bbox, self._last_shape_bbox)
                        if iou_now < self.shape_change_iou:
                            self._shape_change_frames += 1
                        else:
                            self._shape_change_frames = 0
                        # movement check
                        cx_prev = (self._last_shape_bbox[0] + self._last_shape_bbox[2]) / 2
                        cy_prev = (self._last_shape_bbox[1] + self._last_shape_bbox[3]) / 2
                        cx_now = (best_bbox[0] + best_bbox[2]) / 2
                        cy_now = (best_bbox[1] + best_bbox[3]) / 2
                        dist = ((cx_now - cx_prev) ** 2 + (cy_now - cy_prev) ** 2) ** 0.5
                        if dist > self.shape_move_dist_px:
                            self._shape_move_frames += 1
                        else:
                            self._shape_move_frames = 0
                    # update movement state and ROI change state
                    self._last_shape_bbox = best_bbox
                    try:
                        x1, y1, x2, y2 = [max(0, int(v)) for v in best_bbox]
                        roi = cv2.cvtColor(frame_resized[y1:y2, x1:x2], cv2.COLOR_BGR2GRAY)
                        roi_small = cv2.resize(roi, (96, 96)) if roi.size > 0 else None
                        if roi_small is not None and roi_small.size > 0:
                            if self._prev_shape_roi_gray is not None and self._prev_shape_roi_gray.shape == roi_small.shape:
                                mae = float(np.mean(np.abs(roi_small.astype(np.float32) - self._prev_shape_roi_gray.astype(np.float32))))
                                if mae > self.roi_change_mae_thresh:
                                    self._roi_change_frames += 1
                                else:
                                    self._roi_change_frames = 0
                            self._prev_shape_roi_gray = roi_small
                    except Exception:
                        self._roi_change_frames = 0
                else:
                    self._shape_change_frames = 0
                    self._shape_move_frames = 0
                    self._roi_change_frames = 0

                # If we already counted and a new plate appears without a full absence, reset state for next count
                if self._freeze_after_count and (
                    self._shape_change_frames >= self.shape_change_thresh or
                    self._shape_move_frames >= self.shape_move_frames_thresh or
                    self._roi_change_frames >= self.roi_change_frames_thresh
                ):
                    # Reset for new plate
                    self._reset_plate_state()

                if plate_detected:
                    if not self._freeze_after_count:
                        # อัปเดต latched defect counts
                        for k, v in defect_counts.items():
                            self._latched_defect_counts[k] = max(self._latched_defect_counts.get(k, 0), int(v))
                        if shapes_found:
                            # remember last stable shape(s) during the presence window
                            self._latched_shapes = set(shapes_found)
                        self._render_latched_defect_counts()

                if plate_detected:
                    self.gate_present_frames += 1; self.gate_absent_frames = 0
                else:
                    self.gate_absent_frames += 1; self.gate_present_frames = 0

                # เมื่อตรวจพบจานครั้งแรก
                if (not self.gate_has_plate) and plate_detected and self.gate_present_frames >= self.gate_present_thresh:
                    self.gate_has_plate = True
                    self.gate_has_counted = False
                    self._plate_counted_already = False
                    self._defect_detected_flag = False
                    self._plate_final_status = None
                    self._current_defect_count = 0
                    self._set_plate_status("pending")
                    self._latched_shapes = set()
                    self._render_latched_defect_counts()

                # อัปเดตสถานะแบบไดนามิกระหว่างที่จานยังอยู่
                if self.gate_has_plate and self.gate_present_frames >= self.gate_present_thresh:
                    # นับจานเพียงครั้งเดียว
                    if not self._plate_counted_already:
                        self.shape_counts["total"] += 1
                        self.total_number_label.configure(text=str(self.shape_counts["total"]))
                        self.lbl_plate_order.configure(text=str(self.shape_counts["total"]))
                        self._plate_counted_already = True
                    
                    # อัปเดตการนับรูปทรง (เฉพาะครั้งแรกที่เจอ)
                    shapes_to_use = self._latched_shapes if len(self._latched_shapes) > 0 else shapes_found
                    for shp in shapes_to_use:
                        if shp in self.shape_counts:
                            # ตรวจสอบว่าเคยนับรูปทรงนี้ไปแล้วหรือยัง
                            if not hasattr(self, '_counted_shapes'):
                                self._counted_shapes = set()
                            if shp not in self._counted_shapes:
                                self.shape_counts[shp] += 1
                                self._counted_shapes.add(shp)
                    
                    self.lbl_heart.configure(text=str(self.shape_counts["heart"]))
                    self.lbl_rect.configure(text=str(self.shape_counts["rectangle"]))
                    self.lbl_circle.configure(text=str(self.shape_counts["circle"]))

                    # ตรวจสอบตำหนิแบบไดนามิก
                    latched_defect_total = int(sum(self._latched_defect_counts.values()))
                    current_defect_total = int(sum(defect_counts.values()))
                    has_defects_now = latched_defect_total > 0 or current_defect_total > 0 or len(defect_names) > 0
                    
                    # อัปเดตสถานะการตรวจพบตำหนิ
                    if has_defects_now and not self._defect_detected_flag:
                        # เจอตำหนิครั้งแรก - ตีตราถาวร
                        self._defect_detected_flag = True
                        self._current_defect_count = max(latched_defect_total, current_defect_total, len(defect_names))
                        self._plate_final_status = "defect"
                        
                        # บันทึกข้อมูลทันทีเมื่อเจอตำหนิ
                        if not self.gate_has_counted:
                            self._save_detection_record(annotated, defect_names, shapes_to_use)
                            self.gate_has_counted = True
                        
                        self._set_plate_status("counted", self._current_defect_count)
                        self._log_with_emoji("warning", f"พบตำหนิ {self._current_defect_count} จุด ในจานที่ {self.plate_id_counter-1}")
                        
                    elif has_defects_now and self._defect_detected_flag:
                        # อัปเดตจำนวนตำหนิถ้ามีมากกว่าเดิม
                        new_defect_count = max(latched_defect_total, current_defect_total, len(defect_names))
                        if new_defect_count > self._current_defect_count:
                            self._current_defect_count = new_defect_count
                            self._set_plate_status("counted", self._current_defect_count)
                            
                    elif not has_defects_now and not self._defect_detected_flag:
                        # ยังไม่เจอตำหนิ - สถานะผ่านชั่วคราว
                        if self._plate_final_status != "defect":
                            self._plate_final_status = "pass"
                            self._set_plate_status("counted", 0)
                    
                    # ถ้ายังไม่เคยบันทึกและจานกำลังจะออกไป ให้บันทึก
                    if not self.gate_has_counted and self.gate_absent_frames > 0:
                        final_defects = defect_names if self._defect_detected_flag else set()
                        self._save_detection_record(annotated, final_defects, shapes_to_use)
                        self.gate_has_counted = True
                        
                        if not self._defect_detected_flag:
                            self._log_with_emoji("success", f"จานที่ {self.plate_id_counter-1} ผ่านการตรวจสอบ")

                # เมื่อจานออกไปจากระบบ
                if self.gate_has_plate and self.gate_absent_frames >= self.gate_absent_thresh:
                    # บันทึกข้อมูลครั้งสุดท้ายก่อนรีเซ็ต (ถ้ายังไม่ได้บันทึก)
                    if not self.gate_has_counted:
                        shapes_to_use = self._latched_shapes if len(self._latched_shapes) > 0 else shapes_found
                        final_defects = set()
                        if self._defect_detected_flag:
                            # ใช้ defect_names ล่าสุดหรือ latched
                            for k, v in self._latched_defect_counts.items():
                                if v > 0:
                                    final_defects.add(k)
                        
                        self._save_detection_record(annotated, final_defects, shapes_to_use)
                        
                        if self._defect_detected_flag:
                            self._log_with_emoji("warning", f"พบตำหนิ {self._current_defect_count} จุด ในจานที่ {self.plate_id_counter-1}")
                        else:
                            self._log_with_emoji("success", f"จานที่ {self.plate_id_counter-1} ผ่านการตรวจสอบ")
                    
                    self._reset_plate_state()

            except Exception as e:
                print(f"Inference error: {e}")
                frame_to_show = frame_resized
       
        else:
            frame_to_show = frame_resized

        try:
            frame_rgb = cv2.cvtColor(frame_to_show, cv2.COLOR_BGR2RGB)
            pil_img = Image.fromarray(frame_rgb)
            imgtk = ImageTk.PhotoImage(pil_img)
            self.camera_label.configure(image=imgtk, text="")
            self.camera_label.image = imgtk
        except Exception as e:
            print(f"Camera display error: {e}")

        self.app.after(30, self.update_camera)

    def _reset_plate_state(self):
        """รีเซ็ตสถานะสำหรับจานใหม่"""
        self.gate_has_plate = False
        self.gate_has_counted = False
        self._plate_counted_already = False
        self._defect_detected_flag = False
        self._plate_final_status = None
        self._current_defect_count = 0
        self._set_plate_status("pending")
        self._latched_shapes = set()
        self._latched_defect_counts = {"crack": 0, "hole": 0, "bulge": 0, "burn": 0}
        self._reset_defect_table()
        self._render_latched_defect_counts()
        self._freeze_after_count = False
        self.gate_present_frames = 0
        self.gate_absent_frames = 0
        self._prev_shape_roi_gray = None
        if hasattr(self, '_counted_shapes'):
            self._counted_shapes = set()

    # ---------------- Events ----------------
    def toggle_data_collection(self):
        if not self.is_collecting_data:
            # ตรวจสอบข้อมูลเดิมและอัปเดต lot_id และ plate_id_counter
            self._check_and_continue_session()
            self.lot_id = self.generate_lot_id()
            self._update_lot_label()
            
            self.is_collecting_data = True
            self.session_meta.setdefault("start_time", datetime.now().strftime("%H:%M:%S"))
            self.toggle_button.configure(text="หยุด", fg_color=self.COLOR_DANGER, hover_color=self.COLOR_DANGER_HOVER)
            self.gate_has_plate = self.gate_has_counted = False
            self.gate_present_frames = self.gate_absent_frames = 0
            self._set_plate_status("pending")
            self._reset_defect_table(); self._render_latched_defect_counts()
            self._ensure_session_files(); self._update_excel_session_times()
            self._log_with_emoji("success", f"เริ่มการตรวจสอบจานใบไม้ (ชุด: {self.lot_id})")
        else:
            self.show_stop_confirm_dialog()

    def show_stop_confirm_dialog(self):
        px = self.px
        dlg = ctk.CTkToplevel(self.app)
        dlg.title("ยืนยันการหยุด"); dlg.geometry(f"{px(380)}x{px(190)}"); dlg.resizable(False, False)
        dlg.transient(self.app); dlg.grab_set()
        warning_emoji = self._get_colored_emoji(":warning:")
        ctk.CTkLabel(dlg, text=f"{warning_emoji} คุณยืนยันจะหยุดตรวจและบันทึกรายงานหรือไม่?", font=self.F(16, True))\
            .place(x=px(190), y=px(55), anchor="center")
        ctk.CTkButton(
            dlg, width=px(140), height=px(40), text="Submit (ยืนยัน)", font=self.F(14, True),
            fg_color=self.COLOR_PRIMARY, hover_color=self.COLOR_PRIMARY_HOVER,
            command=lambda: (dlg.destroy(), self.stop_and_finalize())
        ).place(x=px(40), y=px(110))
        ctk.CTkButton(
            dlg, width=px(140), height=px(40), text="ตรวจต่อ", font=self.F(14, True),
            fg_color=self.COLOR_NEUTRAL, hover_color=self.COLOR_NEUTRAL_HOVER,
            command=dlg.destroy
        ).place(x=px(200), y=px(110))

    def stop_and_finalize(self):
        self.is_collecting_data = False
        self._log_with_emoji("info", "หยุดการตรวจสอบจานใบไม้")
        try:
            self.toggle_button.configure(text="เริ่ม", fg_color=self.COLOR_PRIMARY, hover_color=self.COLOR_PRIMARY_HOVER)
        except Exception:
            pass

        if self._session_stamp:
            self._update_excel_session_times()
            meta = {
                "report_title": f"รายงานการตรวจจานใบไม้ วันที่ {self._title_date(datetime.now())}",
                "lot_id": self.lot_id,
                "session": {
                    "start_time": self.session_meta.get("start_time"),
                    "end_time": datetime.now().strftime("%H:%M:%S")
                }
            }
            self._firebase_put(f"sessions/{self.firebase_session_key}/meta", meta)
            self._log_with_emoji("success", "บันทึกรายงานเรียบร้อย")

        self._reset_all_and_next_lot()

    def _reset_all_and_next_lot(self):
        self.shape_counts = {"heart": 0, "rectangle": 0, "circle": 0, "total": 0}
        try:
            self.lbl_heart.configure(text="0")
            self.lbl_rect.configure(text="0")
            self.lbl_circle.configure(text="0")
            self.total_number_label.configure(text="0")
            self.lbl_plate_order.configure(text="0")
            self.lbl_defect_count.configure(text="ยังไม่พบ", text_color="#888888")
            self.lbl_plate_status.configure(text="ยังไม่ได้ตรวจ", text_color="#888888")
        except Exception:
            pass

        self.session_rows = []
        self.session_meta = {}
        self.plate_id_counter = 1
        self._auto_xlsx_path = None
        self._auto_json_path = None
        self._session_stamp = None
        self.firebase_session_key = None

        self.gate_has_plate = self.gate_has_counted = False
        self.gate_present_frames = self.gate_absent_frames = 0
        self._set_plate_status("pending")
        self._reset_defect_table(); self._render_latched_defect_counts()

        self._increment_lot_id()

    def update_header_time(self):
        self.header_time_label.configure(text=datetime.now().strftime("%H:%M:%S"))
        self.app.after(1000, self.update_header_time)

    def on_closing(self):
        if self.is_collecting_data:
            warning_emoji = self._get_colored_emoji(":warning:")
            if not messagebox.askyesno("ออกจากโปรแกรม", f"{warning_emoji} กำลังตรวจจับอยู่ ต้องการออกทันทีหรือไม่?"):
                return
            self.stop_and_finalize()
        self.stop_camera()
        self.app.destroy()

    # ---------------- Excel Viewer launchers ----------------
    def open_excel_viewer(self):
        ExcelViewerDialog(self.app, self.save_root, title="Excel Viewer")

    def open_weekly_viewer(self):
        # สร้าง/อัปเดตรายงานรายสัปดาห์ของสัปดาห์ปัจจุบันก่อนเปิด
        start_d, end_d = self._get_week_range_mon_sun(date.today())
        self._ensure_weekly_report(start_d, end_d)
        WeeklyExcelViewerDialog(self.app, self.save_root, title="Excel Viewer - Weekly")

    # ---------------- Weekly report helpers ----------------
    @staticmethod
    def _get_week_range_mon_sun(d: date):
        """คืนค่า (วันจันทร์, วันอาทิตย์) ของสัปดาห์ที่ d อยู่"""
        mon = d - timedelta(days=(d.weekday()))       # Monday
        sun = mon + timedelta(days=6)                 # Sunday
        return mon, sun

    @staticmethod
    def _thai_weekday_name(d: date):
        names = ["วันจันทร์", "วันอังคาร", "วันพุธ", "วันพฤหัสบดี", "วันศุกร์", "วันเสาร์", "วันอาทิตย์"]
        return names[d.weekday()]

    @staticmethod
    def _parse_thai_date(s: str):
        # รูปแบบ dd/mm/BBBB (พ.ศ.)
        try:
            dd, mm, bb = s.split("/")
            return date(int(bb) - 543, int(mm), int(dd))
        except Exception:
            return None

    def _ensure_weekly_report(self, start_d: date, end_d: date):
        fname = f"Weekly_{start_d.strftime('%Y%m%d')}-{end_d.strftime('%Y%m%d')}.xlsx"
        fpath = os.path.join(self.save_root, fname)
        try:
            self._build_weekly_excel(fpath, start_d, end_d)
        except Exception as e:
            messagebox.showerror("Weekly Report", f"สร้างรายงานรายสัปดาห์ไม่สำเร็จ:\n{e}")

    def _build_weekly_excel(self, path, start_d: date, end_d: date):
        """อ่านไฟล์รายวันทั้งหมดในโฟลเดอร์ -> สรุปเป็นรายสัปดาห์ [จันทร์-อาทิตย์]
           'หัววัน' จะเป็นชื่อวันอย่างเดียว (ไม่มีตัวเลขวันที่)
        """
        stats = {}
        defect_counter = defaultdict(Counter)

        for rp in glob.glob(os.path.join(self.save_root, "Report_*.xlsx")):
            wb = openpyxl.load_workbook(rp, data_only=True)
            sh = wb.active
            head_r = 1
            for r in range(1, min(30, sh.max_row) + 1):
                if (sh.cell(row=r, column=1).value or "") == "วันที่":
                    head_r = r
                    break
            for r in range(head_r + 1, sh.max_row + 1):
                d_th = sh.cell(row=r, column=1).value
                lot = sh.cell(row=r, column=4).value
                defects = sh.cell(row=r, column=6).value
                if not d_th or not lot:
                    continue
                gd = self._parse_thai_date(str(d_th))
                if gd is None:
                    continue
                if not (start_d <= gd <= end_d):
                    continue

                key = (gd, str(lot))
                st = stats.setdefault(key, {"total": 0, "ok": 0, "ng": 0})
                st["total"] += 1
                if defects and str(defects).strip() not in ("-", ""):
                    st["ng"] += 1
                    for token in [t.strip() for t in str(defects).split("/") if t.strip()]:
                        defect_counter[key][token] += 1
                else:
                    st["ok"] += 1
            wb.close()

        wb = openpyxl.Workbook()
        sh = wb.active
        sh.title = "Weekly"

        title = f"รายงานการตรวจจานใบไม้ ช่วงวันที่ {start_d.strftime('%d/%m/%y')} - {end_d.strftime('%d/%m/%y')}"
        sh.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        sh["A1"] = title
        sh["A1"].font = Font(size=14, bold=True)
        sh["A1"].alignment = Alignment(horizontal="center")

        headers = ["วันที่", "รหัสชุด", "จำนวนจานทั้งหมด", "ไม่มีตำหนิ", "มีตำหนิ", "ตำหนิที่พบบ่อยที่สุด", "หมายเหตุ"]
        sh.append([])
        sh.append(headers)
        head_row = 3
        head_font = Font(bold=True)
        fill = PatternFill("solid", fgColor="E4DFDA")
        align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="B7B7B7")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for c in range(1, len(headers) + 1):
            cell = sh.cell(row=head_row, column=c)
            cell.font = head_font; cell.fill = fill; cell.alignment = align; cell.border = border
            sh.column_dimensions[get_column_letter(c)].width = 22

        r = head_row + 1
        total_all, ok_all, ng_all = 0, 0, 0
        cur = start_d
        while cur <= end_d:
            # ---- แถวหัววัน (ชื่อวันอย่างเดียว) ----
            day_bar = self._thai_weekday_name(cur)
            sh.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            cell = sh.cell(row=r, column=1, value=day_bar)
            cell.font = Font(bold=True, color="333333")
            cell.fill = PatternFill("solid", fgColor="E6E6E6")
            cell.alignment = Alignment(horizontal="left")
            r += 1

            day_keys = sorted([k for k in stats.keys() if k[0] == cur], key=lambda x: x[1])
            for key in day_keys:
                gd, lot = key
                st = stats[key]
                total_all += st["total"]; ok_all += st["ok"]; ng_all += st["ng"]
                most = "-"
                if defect_counter[key]:
                    most = defect_counter[key].most_common(1)[0][0]

                sh.append([gd.strftime("%d/%m/%Y"), lot, st["total"], st["ok"], st["ng"], most, ""])
                r += 1

            cur += timedelta(days=1)

        sh.append(["รวมทั้งหมด", "", total_all, ok_all, ng_all, "", ""])
        sum_row = sh.max_row
        sh.cell(row=sum_row, column=1).font = Font(bold=True)
        sh.cell(row=sum_row, column=3).font = Font(bold=True)
        sh.cell(row=sum_row, column=4).font = Font(bold=True)
        sh.cell(row=sum_row, column=5).font = Font(bold=True)

        wb.save(path)
        wb.close()

    # ---------------- Run ----------------
    def run(self):
        self.app.mainloop()


# ================================
# Boot
# ================================
if __name__ == "__main__":
    try:
        signal.signal(signal.SIGINT, signal.SIG_IGN)
    except Exception:
        pass

    print("🚀 กำลังเริ่มต้นโปรแกรมตรวจจับรอยตำหนิบนจานใบไม้...")

    app = LeafPlateDetectionApp()
    app.initialize_data()
    app.setup_app()
    app.setup_fonts()
    app.setup_camera()
    app.setup_model()
    app.create_widgets()
    app.start_camera()

    print("✅ โปรแกรมพร้อมใช้งาน!")

    try:
        app.run()
    except KeyboardInterrupt:
        print("🛑 Interrupted by user. Exiting cleanly...")
